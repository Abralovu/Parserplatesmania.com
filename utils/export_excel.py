import os
from typing import Optional

import aiosqlite
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config.settings import DB_PATH, OUTPUT_DIR
from utils.logger import get_logger

logger = get_logger(__name__)

_HEADER_FILL = PatternFill("solid", fgColor="1F3864")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
_ROW_HEIGHT = 70
_COL_WIDTHS = [20, 10, 14, 8, 28, 18, 16, 55, 12]
_HEADERS = ["Фото", "ID", "Номер", "Страна", "Регион", "Марка", "Модель", "Ссылка на фото", "Дата"]


async def export_excel_with_photos(
    country: Optional[str] = None,
    region: Optional[str] = None,
    car_brand: Optional[str] = None,
) -> Optional[str]:
    """
    Экспортирует данные в Excel с фото.
    Фильтры: country, region, car_brand.
    Возвращает путь к файлу или None.
    """
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    rows = await _fetch_rows(country, region, car_brand)
    if not rows:
        logger.warning("No data to export")
        return None

    suffix = _build_suffix(country, region, car_brand)
    output_path = os.path.abspath(os.path.join(OUTPUT_DIR, f"plates{suffix}.xlsx"))

    wb = Workbook()
    ws = wb.active
    ws.title = "Plates"

    _write_headers(ws)

    for row_idx, record in enumerate(rows, 2):
        _write_row(ws, row_idx, record)

    for col, width in enumerate(_COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = "C2"
    wb.save(output_path)

    logger.info(f"Excel exported: {output_path} ({len(rows)} records)")
    return output_path


# ─── Приватные функции ────────────────────────────────────────────────────────

async def _fetch_rows(
    country: Optional[str],
    region: Optional[str],
    car_brand: Optional[str],
) -> list:
    """Выбирает строки из БД с фильтрами."""
    conditions = []
    params = []

    if country:
        conditions.append("country=?")
        params.append(country)
    if region:
        conditions.append("region LIKE ?")
        params.append(f"%{region}%")
    if car_brand:
        conditions.append("car_brand LIKE ?")
        params.append(f"%{car_brand}%")

    where = f"WHERE {' AND '.join(conditions)}" if conditions else ""
    sql = f"SELECT * FROM plates {where} ORDER BY country, region, car_brand, plate_id"

    async with aiosqlite.connect(DB_PATH) as db:
        db.row_factory = aiosqlite.Row
        cursor = await db.execute(sql, params)
        return await cursor.fetchall()


def _build_suffix(
    country: Optional[str],
    region: Optional[str],
    car_brand: Optional[str],
) -> str:
    """Строит суффикс имени файла из фильтров."""
    parts = []
    if country:
        parts.append(country)
    if region:
        parts.append(region.replace(" ", "_")[:20])
    if car_brand:
        parts.append(car_brand.replace(" ", "_")[:15])
    return f"_{'_'.join(parts)}" if parts else "_all"


def _write_headers(ws) -> None:
    """Записывает строку заголовков."""
    for col, header in enumerate(_HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _THIN_BORDER
    ws.row_dimensions[1].height = 25


def _write_row(ws, row_idx: int, record) -> None:
    """Записывает одну строку данных."""
    ws.row_dimensions[row_idx].height = _ROW_HEIGHT
    row_fill = PatternFill("solid", fgColor="EBF0FA" if row_idx % 2 == 0 else "FFFFFF")

    # Фото в колонке A
    _insert_photo(ws, row_idx, record["local_path"], record["plate_id"])

    # Данные начиная с колонки B
    values = [
        record["plate_id"],
        record["plate_number"],
        record["country"].upper() if record["country"] else "—",
        record["region"] or "—",
        record["car_brand"] or "—",
        record["car_model"] or "—",
        record["photo_url"] or "—",
        record["scraped_at"][:10] if record["scraped_at"] else "—",
    ]

    for col_idx, value in enumerate(values, 2):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.fill = row_fill
        cell.alignment = Alignment(vertical="center", horizontal="left")
        cell.border = _THIN_BORDER


def _insert_photo(ws, row_idx: int, local_path: Optional[str], plate_id: int) -> None:
    """Вставляет фото в ячейку или текст-заглушку."""
    if not local_path:
        ws.cell(row=row_idx, column=1, value="нет фото")
        return

    abs_path = os.path.abspath(local_path)
    if not os.path.exists(abs_path):
        ws.cell(row=row_idx, column=1, value="файл не найден")
        return

    try:
        img = XLImage(abs_path)
        img.height = 85
        img.width = 125
        img.anchor = f"A{row_idx}"
        ws.add_image(img)
    except Exception as e:
        ws.cell(row=row_idx, column=1, value="ошибка фото")
        logger.warning(f"Image error id={plate_id}: {e}")